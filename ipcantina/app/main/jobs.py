from app.main.email import send_daily_summary_email
from datetime import date
from tabulate import tabulate
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font
import os
from flask import current_app

# from app.models import Meal, Order
# from app.main.utils import DateUtils

from db.models import Meal, Order
from db.utils import DateUtils
from db.config import config


def create_daily_order_sheet(dt):
    meals = Meal.query.filter(Meal.date == dt, Meal.label != 'S').order_by(Meal.label).all()
    summary = []
    for meal in meals:
        amount = Order.query.with_entities(func.count()).filter(Order.meal_id == meal.id).scalar()
        summary.append((meal, amount))

    date_str = DateUtils.to_string(dt)
    wb = Workbook()
    ws = wb.active

    bold = Font(bold=True)
    ws['B1'] = "Denné menu " + date_str; ws['B1'].font = bold
    ws['C1'] = "online"; ws['C1'].font = bold
    ws['D1'] = "voľný predaj"; ws['D1'].font = bold
    ws['E1'] = "spolu"; ws['E1'].font = bold

    for i, item in enumerate(summary):
        row = 3+i
        ws.cell(row=row, column=1, value=item[0].label)
        ws.cell(row=row, column=2, value=item[0].description)
        ws.cell(row=row, column=3, value=item[1])
        sum = "=SUM(C" + str(row) + ", D" + str(row) + ")"
        ws.cell(row=row, column=5, value=sum)

    widths = {'A': 4, 'B': 50, 'C': 10, 'D': 10, 'E': 10}

    for k, v in widths.items():
        ws.column_dimensions[k].width = v

    filename = os.path.join(current_app.config['ATTACHMENTS_DIR_PATH'], "IP_objednavka_" + date_str + ".xlsx")
    wb.save(filename)

    return filename


headers = ['Objednávka', 'Meno', 'Tel. číslo', 'So sebou', 'Cena']


def create_order_list_file(dt):
    orders, descriptions = Order.get_orders_for_day(dt)
    orders = orders.order_by(Order.user_id).all()
    data = []
    counts = {'A': 0, 'B': 0, 'C': 0}
    for order in orders:
        meal = Meal.query.get(order.meal_id)
        price = '%.2f €' % (meal.price + config['MEAL_BOX_PRICE'] if order.take_away else meal.price)
        take_away = 'áno' if order.take_away else ''
        name = ' '.join([order.customer.surname, order.customer.first_name])
        data.append([meal.label, name, order.customer.phone, take_away, price])
        counts[meal.label] += 1

    table = tabulate(data, headers=headers)
    # print(table)
    date_str = DateUtils.to_string(dt)
    filename = os.path.join(current_app.config['ATTACHMENTS_DIR_PATH'], "IP_cantina_zoznam_" + date_str + ".txt")
    # filename = '\\'.join([current_app.config['ATTACHMENTS_DIR_PATH'], "IP_cantina_zoznam_" + date_str + ".txt"])
    from random import uniform
    from time import sleep
    sleep(uniform(0., 5.))
    with open(filename, "w", encoding='utf-8') as f:
        f.write("%s %s - ONLINE OBJEDNÁVKY\n\n" % (DateUtils.to_string(dt), DateUtils.svk_weekday_from_int(dt.weekday())))
        f.write(table)
        f.write('\n\n')
        f.write("Spolu:\n")
        for i, k in enumerate(counts.keys()):
            f.write("%dx %s: %s\n" % (counts[k], k, descriptions[i]))

    return filename


def send_daily_summary(app):
    with app.app_context():
        app.logger.info("Daily emailing job has started")
        today = date.today()
        if today.weekday() == 5 or today.weekday() == 6:
            return
        dt = DateUtils.next_working_day(today)
        order_list = create_order_list_file(dt)
        order_sheet = create_daily_order_sheet(dt)
        app.logger.info("Attachments created")
        send_daily_summary_email(dt, order_list, order_sheet)
        app.logger.info("Email was successfully sent")

