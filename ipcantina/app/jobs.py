from app.email import send_daily_summary_email
from datetime import date, timedelta
from tabulate import tabulate
from app.models import Meal, Order
from app.utils import DateUtils
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font
import os
from app import app

headers = ['Priezvisko', 'Tel. číslo', 'Objednávka']


def create_daily_order_sheet(dt):
    meals = Meal.query.filter(Meal.date == dt, Meal.label != 'S').order_by(Meal.label).all()
    summary = []
    for meal in meals:
        amount = Order.query.with_entities(func.count()).filter(Order.meal_id == meal.id).scalar()
        summary.append((meal, amount))

    date_str = DateUtils.to_string(dt)
    wb = Workbook()
    ws = wb.active

    bold = Font(bold=True)
    ws['B1'] = "Denné menu " + date_str; ws['B1'].font = bold
    ws['C1'] = "online"; ws['C1'].font = bold
    ws['D1'] = "voľný predaj"; ws['D1'].font = bold
    ws['E1'] = "dokopy"; ws['E1'].font = bold

    for i, item in enumerate(summary):
        row = 3+i
        ws.cell(row=row, column=1, value=item[0].label)
        ws.cell(row=row, column=2, value=item[0].description)
        ws.cell(row=row, column=3, value=item[1])
        sum = "=SUM(C" + str(row) + ", D" + str(row) + ")"
        ws.cell(row=row, column=5, value=sum)

    # dims = {} TODO BUG HERE
    # for row in ws.rows:
    #     for cell in row:
    #         if cell.value:
    #             dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))

    # for col, value in dims.items():
    #     ws.column_dimensions[col].width = value * 1.2

    filename = os.path.join(app.config['ATTACHMENTS_DIR_PATH'], "IP_objednavka_" + date_str + ".xlsx")
    wb.save(filename)

    return filename


def create_order_list_file(dt):
    orders = Order.get_orders_for_day(dt).all()
    data = []
    for order in orders:
        meal = Meal.query.get(order.meal_id)
        data.append([order.customer.surname, order.customer.phone, meal.label])
    table = tabulate(data, headers=headers)
    # print(table)
    date_str = DateUtils.to_string(dt)
    filename = os.path.join(app.config['ATTACHMENTS_DIR_PATH'], "IP_cantina_zoznam_" + date_str + ".txt")

    with open(filename, "w", encoding='utf-8') as f:
        f.write(DateUtils.to_string(dt) + "\n\n")
        f.write(table)

    return filename


def send_daily_summary():
    dt = DateUtils.next_working_day()

    order_list = create_order_list_file(dt)
    order_sheet = create_daily_order_sheet(dt)

    send_daily_summary_email(dt, order_list, order_sheet)

